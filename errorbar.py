# Author: Yusuke Yoshida
# Created: 7/11/2016
# Comment: Example for handling error bars with openpylx

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.data_source import NumDataSource, NumData, NumVal

def main():
    # Create workbood and select active worksheet
    wb = Workbook()
    ws = wb.active

    # Write labels
    ws.append(["X", "Y", "Error Plus", "Error Minus"])

    # Define sample data
    x = range(1, 6)
    y = [1, 2, 4, 3, 7]
    plus = [0.3, 0.2, 0.5, 0.3, 0.4]
    minus = [0.2, 0.5, 0.8, 0.3, 0.4]

    # Writer sample data to the worksheet
    for i in range(5):
        ws.cell(row=2+i, column=1).value = x[i]
        ws.cell(row=2+i, column=2).value = y[i]
        ws.cell(row=2+i, column=3).value = plus[i]
        ws.cell(row=2+i, column=4).value = minus[i]

    # Define References for X and Y
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=6)
    yvalues = Reference(ws, min_col=2, min_row=2, max_row=6)

    # Define errorbars. errDir: 'x': x direction error, 'y': y direction error
    errorbars = list2errorbars(plus, minus, errDir='y')

    # Create series data and set error bar
    series = SeriesFactory(yvalues, xvalues, title="y direction error")
    series.errBars = errorbars

    # Create chart and set the series data
    chart = get_chart()
    chart.series.append(series)

    # Add chart to the worksheet
    ws.add_chart(chart, "E2")

    # Save workbook
    wb.save("errorbar.xlsx")




def list2errorbars(plus, minus, errDir='y', errValType='cust'):
    "Returns ErrorBar from lists of error values"

    #Convert to list of NumVal
    numvals_plus = [NumVal(i, None, v=x) for i,x in enumerate(plus)]
    numvals_minus = [NumVal(i, None, v=x) for i,x in enumerate(minus)]

    # Convert to NumData
    nd_plus = NumData(pt=numvals_plus)
    nd_minus = NumData(pt=numvals_minus)

    # Convert to NumDataSource
    nds_plus = NumDataSource(numLit=nd_plus)
    nds_minus = NumDataSource(numLit=nd_minus)

    return ErrorBars(plus=nds_plus, minus=nds_minus, errDir=errDir, errValType=errValType)

def get_chart():
    "Returns schatter chart for example"
    chart = ScatterChart()
    chart.height = 10
    chart.width = 15
    chart.style = 2
    chart.x_axis.title = "X"
    chart.y_axis.title = "Y"
    return chart

if __name__=='__main__':
    main()
